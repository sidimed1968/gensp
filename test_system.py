"""
Script de test pour vérifier le système de gestion des logements
"""

import sys
import os

def test_imports():
    """Test des imports des bibliothèques"""
    print("🧪 Test des imports...")
    
    try:
        import pandas as pd
        print("  ✅ pandas")
    except ImportError as e:
        print(f"  ❌ pandas: {e}")
        return False
    
    try:
        import openpyxl
        print("  ✅ openpyxl")
    except ImportError as e:
        print(f"  ❌ openpyxl: {e}")
        return False
    
    try:
        import streamlit as st
        print("  ✅ streamlit")
    except ImportError as e:
        print(f"  ❌ streamlit: {e}")
        return False
    
    try:
        import folium
        print("  ✅ folium")
    except ImportError as e:
        print(f"  ❌ folium: {e}")
        return False
    
    try:
        import plotly
        print("  ✅ plotly")
    except ImportError as e:
        print(f"  ❌ plotly: {e}")
        return False
    
    try:
        import arabic_reshaper
        print("  ✅ arabic_reshaper")
    except ImportError as e:
        print(f"  ❌ arabic_reshaper: {e}")
        return False
    
    try:
        from bidi.algorithm import get_display
        print("  ✅ python-bidi")
    except ImportError as e:
        print(f"  ❌ python-bidi: {e}")
        return False
    
    return True


def test_database_module():
    """Test du module database"""
    print("\n🧪 Test du module database...")
    
    try:
        from database import LogementDatabase
        print("  ✅ Import du module database réussi")
        
        # Créer une instance de test
        db = LogementDatabase(db_path="test_logements.db", excel_path="logements.xlsx")
        print("  ✅ Initialisation de la base de données réussie")
        
        # Tester les méthodes de base
        stats = db.obtenir_statistiques()
        print(f"  ✅ Statistiques obtenues: {stats.get('total', 0)} logements")
        
        # Tester l'import si le fichier existe
        if os.path.exists("logements.xlsx"):
            print("  ℹ️  Fichier Excel trouvé, test d'import...")
            count, message = db.importer_depuis_excel()
            print(f"  ✅ Import testé: {message}")
        else:
            print("  ⚠️  Fichier Excel non trouvé, import non testé")
        
        # Nettoyer
        db.fermer()
        if os.path.exists("test_logements.db"):
            os.remove("test_logements.db")
        
        return True
        
    except Exception as e:
        print(f"  ❌ Erreur: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_excel_file():
    """Test du fichier Excel"""
    print("\n🧪 Test du fichier Excel...")
    
    if not os.path.exists("logements.xlsx"):
        print("  ⚠️  Fichier logements.xlsx non trouvé")
        return False
    
    try:
        import pandas as pd
        import openpyxl
        
        # Tester avec openpyxl
        wb = openpyxl.load_workbook("logements.xlsx", read_only=True, data_only=True)
        print(f"  ✅ Fichier Excel valide")
        print(f"  ℹ️  Feuilles: {wb.sheetnames}")
        
        ws = wb.active
        print(f"  ℹ️  Feuille active: {ws.title}")
        
        # Compter les lignes
        row_count = 0
        for row in ws.iter_rows():
            row_count += 1
        print(f"  ℹ️  Nombre de lignes: {row_count}")
        
        wb.close()
        
        # Tester avec pandas (petite lecture)
        df = pd.read_excel("logements.xlsx", nrows=5)
        print(f"  ✅ Lecture pandas réussie")
        print(f"  ℹ️  Colonnes: {len(df.columns)}")
        
        return True
        
    except Exception as e:
        print(f"  ❌ Erreur: {e}")
        return False


def test_app_structure():
    """Test de la structure de l'application"""
    print("\n🧪 Test de la structure de l'application...")
    
    required_files = [
        'app.py',
        'database.py',
        'requirements.txt',
        'README.md',
        'run_app.sh'
    ]
    
    all_ok = True
    for file in required_files:
        if os.path.exists(file):
            size = os.path.getsize(file)
            print(f"  ✅ {file} ({size} bytes)")
        else:
            print(f"  ❌ {file} manquant")
            all_ok = False
    
    return all_ok


def main():
    """Fonction principale de test"""
    print("=" * 60)
    print("🏘️  TEST DU SYSTÈME DE GESTION DES LOGEMENTS")
    print("=" * 60)
    
    results = []
    
    # Test 1: Imports
    results.append(("Imports des bibliothèques", test_imports()))
    
    # Test 2: Structure
    results.append(("Structure de l'application", test_app_structure()))
    
    # Test 3: Fichier Excel
    results.append(("Fichier Excel", test_excel_file()))
    
    # Test 4: Module database
    results.append(("Module database", test_database_module()))
    
    # Résumé
    print("\n" + "=" * 60)
    print("📊 RÉSUMÉ DES TESTS")
    print("=" * 60)
    
    for test_name, result in results:
        status = "✅ RÉUSSI" if result else "❌ ÉCHOUÉ"
        print(f"{status} : {test_name}")
    
    all_passed = all(result for _, result in results)
    
    print("\n" + "=" * 60)
    if all_passed:
        print("🎉 TOUS LES TESTS SONT PASSÉS !")
        print("\n✨ Le système est prêt à être utilisé")
        print("\n🚀 Pour lancer l'application, exécutez:")
        print("   streamlit run app.py")
        print("\n   ou")
        print("   ./run_app.sh")
    else:
        print("⚠️  CERTAINS TESTS ONT ÉCHOUÉ")
        print("\n🔧 Veuillez corriger les erreurs avant de lancer l'application")
    print("=" * 60)
    
    return 0 if all_passed else 1


if __name__ == "__main__":
    sys.exit(main())
